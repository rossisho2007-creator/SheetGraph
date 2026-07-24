from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
import pandas as pd
import os, re, json, uuid
import psycopg2
import psycopg2.extras
from datetime import datetime, date
from io import BytesIO
from functools import wraps

app = Flask(__name__, static_folder='static', static_url_path='/static')
CORS(app)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-only-not-secure-change-me')
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

USERS = {
    'hr': {
        'password': os.environ.get('HR_PASSWORD', 'change-me-hr'),
        'role': 'hr', 'name': 'HR Officer', 'branch': 'Head Office'
    },
    'manager': {
        'password': os.environ.get('MANAGER_PASSWORD', 'change-me-manager'),
        'role': 'manager', 'name': 'Branch Manager', 'branch': 'Jakarta'
    },
    'staff': {
        'password': os.environ.get('STAFF_PASSWORD', 'change-me-staff'),
        'role': 'staff', 'name': 'Staff', 'branch': 'Cabang'
    },
}

def login_required(f):
    @wraps(f)
    def d(*a, **k):
        if 'user' not in session: return redirect(url_for('login'))
        return f(*a, **k)
    return d

DATABASE_URL = os.environ.get('DATABASE_URL')

class Row:
    """Mimics sqlite3.Row exactly: supports BOTH row[0] positional access
    AND row['column'] named access, plus dict(row) conversion -- so no
    query-calling code anywhere in the app needs to know the DB changed."""
    def __init__(self, keys, values):
        self._keys = keys
        self._values = values

    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        return self._values[self._keys.index(key)]

    def keys(self):
        return self._keys

    def __iter__(self):
        return iter(self._values)

    def __repr__(self):
        return repr(dict(zip(self._keys, self._values)))

class PGCursor:
    def __init__(self, cursor):
        self._cursor = cursor
        self._keys = [d[0] for d in cursor.description] if cursor.description else []

    def fetchone(self):
        row = self._cursor.fetchone()
        return Row(self._keys, row) if row is not None else None

    def fetchall(self):
        return [Row(self._keys, row) for row in self._cursor.fetchall()]

    def __iter__(self):
        for row in self._cursor:
            yield Row(self._keys, row)

class PGConn:
    """Wraps a psycopg2 connection so the rest of the app can keep using
    sqlite3-style '?' placeholders and .execute(...).fetchone()/.fetchall()
    without every query in the app being rewritten individually."""
    def __init__(self):
        self._conn = psycopg2.connect(DATABASE_URL)

    def execute(self, query, params=()):
        pg_query = query.replace('?', '%s')
        cur = self._conn.cursor()
        cur.execute(pg_query, params)
        return PGCursor(cur)

    def commit(self):
        self._conn.commit()

    def close(self):
        self._conn.close()

def get_db():
    return PGConn()

def get_setting(key, default=None):
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row['value'] if row else default

def get_setting_full(key):
    conn = get_db()
    row = conn.execute("SELECT value, updated_by, updated_at FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return dict(row) if row else None

def set_setting(key, value, updated_by):
    conn = get_db()
    conn.execute("""INSERT INTO settings (key, value, updated_by, updated_at) VALUES (?,?,?,?)
                     ON CONFLICT(key) DO UPDATE SET value=excluded.value,
                     updated_by=excluded.updated_by, updated_at=excluded.updated_at""",
                 (key, value, updated_by, datetime.now()))
    conn.commit()
    conn.close()

def init_db():
    conn = get_db()
    conn.execute('''CREATE TABLE IF NOT EXISTS submissions (
        id SERIAL PRIMARY KEY,
        application_id TEXT,
        npk TEXT, nama_lengkap TEXT, employee_type TEXT,
        jabatan_gol TEXT, departemen_cabang TEXT, cabang TEXT,
        type_motor TEXT, no_rangka TEXT, no_mesin TEXT, bpkb TEXT,
        loan_amount DOUBLE PRECISION, down_payment DOUBLE PRECISION, tenure_months INTEGER,
        interest_rate DOUBLE PRECISION, tanggal_mulai TEXT,
        principal DOUBLE PRECISION, total_interest DOUBLE PRECISION,
        monthly_installment DOUBLE PRECISION, outstanding_balance DOUBLE PRECISION,
        remarks TEXT, signature TEXT,
        email TEXT, phone TEXT,
        status_approval TEXT DEFAULT 'Pending',
        rejection_reason TEXT,
        submitted_by TEXT, document_source TEXT,
        submitted_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        approved_date TIMESTAMP,
        insurance_amount DOUBLE PRECISION
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS employee_loans (
        npk TEXT PRIMARY KEY, nama_lengkap TEXT,
        email TEXT, phone TEXT,
        active_loan_count INTEGER DEFAULT 0,
        total_loans_ever INTEGER DEFAULT 0, last_loan_date TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT,
        updated_by TEXT,
        updated_at TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS loan_schedule (
        id SERIAL PRIMARY KEY,
        submission_id INTEGER NOT NULL,
        month_number INTEGER NOT NULL,
        due_date TEXT NOT NULL,
        expected_amount DOUBLE PRECISION,
        balance_after DOUBLE PRECISION,
        is_paid INTEGER DEFAULT 0,
        paid_date TIMESTAMP
    )''')
    conn.commit(); conn.close()

init_db()

KNOWN_DEPARTMENTS = ['AR Management', 'Finance', 'Sales', 'Marketing', 'Operation', 'HR Service', 'Collection', 'Credit', 'Service']
KNOWN_BRANCHES = ['Jakarta Pusat','Jakarta Selatan','Jakarta Utara','Bandung','Surabaya','Medan','Makassar','Denpasar','Palembang','Balikpapan','Batam','Yogyakarta','Semarang','Malang','Bekasi','Tangerang','Depok','Bogor','Padang','Pekanbaru','Samarinda','Banjarmasin','Manado','Lampung','Jambi','Bengkulu','Cirebon','Serang','Karawang','Kediri','Jember','Pontianak','Kendari','Palu','Banda Aceh','Duri','Kelapa Gading','BSD City','Tegal']

def add_months(d, months):
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    days_in_month = [31, 29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28,
                      31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(d.day, days_in_month[month - 1])
    return date(year, month, day)

_ocr_reader = None

def get_ocr_reader():
    global _ocr_reader
    if _ocr_reader is None: 
        import easyocr
        _ocr_reader = easyocr.Reader(['id', 'en'])
    return _ocr_reader

def do_ocr(path):
    try:
        reader = get_ocr_reader()
        results = reader.readtext(path)
        return ' '.join([r[1] for r in results]).strip()
    except Exception as e:
        print(f"OCR failed: {e}")
        return ''

NEXT_LABEL = r'(?:No\.?\s*Pokok|NPK|NOPEK|NIK|Jabatan\s*/?\s*Gol|Departemen|Divisi|Cabang|Bagian|Tgl\s*Masuk|Tgl\s*Pengangkatan|Kantor|Lokasi|No\.?\s*Rangka|Rangka|No\.?\s*Mesin|Mesin|BPKB|$)'

def smart_parse(text):
    if not text: return {}, {}, text
    result, conf = {}, {}
    tests = [
        ('npk', r'(?:NPK|NOPEK|No\.?\s*Pokok(?:\s*Karyawan)?|NIK)\s*:?\s*(\d{3,})'),
        ('nama_lengkap', r'(?:Nama|Name)\s*:?\s*([A-Za-z\s\.]{3,60}?)(?=\s*' + NEXT_LABEL + r')'),
        ('jabatan_gol', r'Jabatan\s*/?\s*Gol\s*:?\s*([A-Za-z0-9\s\/\-\.]{2,30}?)(?=\s*' + NEXT_LABEL + r')'),
        ('departemen_cabang', r'Departemen\s*/?\s*Divisi\s*/?\s*Cabang\s*:?\s*([A-Za-z0-9\s\/\-\.]{3,60}?)(?=\s*' + NEXT_LABEL + r')'),
        ('no_rangka', r'(?:No\.?\s*Rangka|Rangka)\s*:?\s*([A-Za-z0-9\-]{5,30}?)(?=\s*' + NEXT_LABEL + r')'),
        ('no_mesin', r'(?:No\.?\s*Mesin|Mesin)\s*:?\s*([A-Za-z0-9\-]{5,30}?)(?=\s*' + NEXT_LABEL + r')'),
        ('bpkb', r'(?:BPKB)\s*:?\s*([A-Za-z0-9\-]{3,30}?)(?=\s*' + NEXT_LABEL + r')'),
    ]
    for field, pat in tests:
        try:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                v = m.group(1).strip()
                if field == 'npk':
                    v = re.sub(r'[^0-9]', '', v)
                    if len(v) < 3: continue
                if v:
                    result[field] = v; conf[field] = 90
        except: pass
    for known in KNOWN_BRANCHES:
        if known.lower() in text.lower():
            result['cabang'] = known; conf['cabang'] = 85; break
    for known in KNOWN_DEPARTMENTS:
        if known.lower() in text.lower():
            result['departemen_cabang'] = known; conf['departemen_cabang'] = 85; break
    if 'operasional' in text.lower() or 'surveyor' in text.lower() or 'fro' in text.lower():
        result['employee_type'] = 'field'
    return result, conf, text

def calculate(data):
    try:
        loan = float(data.get('loan_amount', 16000000))
        dp = float(data.get('down_payment', 0))
        tenor = int(data.get('tenure_months', 36))
        if data.get('employee_type') == 'office':
            rate = float(get_setting('office_interest_rate', '0.05'))
        else:
            rate = 0.05
        if loan <= 0 or tenor <= 0: return {}
        p = loan - dp; ti = p * rate * (tenor/12); m = (p + ti) / tenor
        return {'principal': round(p), 'total_interest': round(ti), 'monthly_installment': round(m), 'outstanding_balance': round(p+ti), 'interest_rate': rate}
    except: return {}

def generate_excel(export_type='full'):
    conn = get_db()
    if export_type == 'approved': df = pd.read_sql_query("SELECT * FROM submissions WHERE status_approval='Approved'", conn)
    elif export_type == 'pending': df = pd.read_sql_query("SELECT * FROM submissions WHERE status_approval='Pending'", conn)
    else: df = pd.read_sql_query("SELECT * FROM submissions", conn)
    conn.close()
    if 'signature' in df.columns:
        df = df.drop(columns=['signature'])
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as w: df.to_excel(w, index=False)
    output.seek(0)
    return output

@app.route('/export-excel/pending')
@login_required
def export_pending():
    if session.get('role') not in ['hr','manager']: return redirect(url_for('dashboard'))
    return send_file(generate_excel('pending'), download_name=f'Pending_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        u = request.form.get('username','').strip().lower()
        p = request.form.get('password','')

        # Path 1: privileged accounts (hr / manager / staff-demo) - real password required
        if u in USERS and p and USERS[u]['password'] == p:
            session['user'] = u
            session['role'] = USERS[u]['role']
            session['name'] = USERS[u]['name']
            session['branch'] = USERS[u]['branch']
            return redirect(url_for('dashboard'))

        # Path 2: regular employee NPK login - no password, role ALWAYS forced
        # to 'staff' server-side. Client-submitted role is never trusted.
        name = request.form.get('name','').strip()
        branch = request.form.get('branch','').strip()
        if u and len(u) >= 2 and not p and name and branch:
            session['user'] = u
            session['role'] = 'staff'
            session['name'] = name
            session['branch'] = branch
            return redirect(url_for('dashboard'))

        flash('Nama pengguna atau kata sandi salah', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    role = session.get('role'); name = session.get('name')
    if role in ['hr','manager']:
        stats = {
            'pending': conn.execute("SELECT COUNT(*) FROM submissions WHERE status_approval='Pending'").fetchone()[0],
            'approved': conn.execute("SELECT COUNT(*) FROM submissions WHERE status_approval='Approved'").fetchone()[0],
            'total': conn.execute("SELECT COUNT(*) FROM submissions").fetchone()[0],
            'active_loans': conn.execute("SELECT COUNT(*) FROM employee_loans WHERE active_loan_count > 0").fetchone()[0],
        }
        recent = [dict(r) for r in conn.execute("SELECT * FROM submissions ORDER BY submitted_date DESC LIMIT 50")]
    else:
        stats = {
            'pending': conn.execute("SELECT COUNT(*) FROM submissions WHERE status_approval='Pending' AND submitted_by=?",(name,)).fetchone()[0],
            'approved': conn.execute("SELECT COUNT(*) FROM submissions WHERE status_approval='Approved' AND submitted_by=?",(name,)).fetchone()[0],
        }
        recent = [dict(r) for r in conn.execute("SELECT * FROM submissions WHERE submitted_by=? ORDER BY submitted_date DESC LIMIT 50",(name,))]
    conn.close()
    office_rate_percent = float(get_setting('office_interest_rate', '0.05')) * 100
    return render_template('dashboard.html', stats=stats, recent=recent, office_rate_percent=office_rate_percent)

@app.route('/api/schedule/<int:submission_id>')
@login_required
def get_schedule(submission_id):
    if session.get('role') not in ['hr','manager']: return jsonify({'status':'error'}), 403
    conn = get_db()
    rows = conn.execute(
        "SELECT month_number, due_date, expected_amount, balance_after, is_paid FROM loan_schedule "
        "WHERE submission_id=? ORDER BY month_number", (submission_id,)
    ).fetchall()
    conn.close()
    if not rows:
        return jsonify({'status':'empty'})
    today = datetime.now().date()
    schedule = []
    for r in rows:
        r = dict(r)
        due = datetime.strptime(r['due_date'], '%Y-%m-%d').date()
        if r['is_paid']:
            status = 'Paid'
        elif due < today:
            status = 'Overdue'
        elif due.year == today.year and due.month == today.month:
            status = 'Due this month'
        else:
            status = 'Upcoming'
        schedule.append({'month': r['month_number'], 'due_date': due.strftime('%d %b %Y'),
                          'amount': r['expected_amount'], 'balance': r['balance_after'], 'status': status})
    return jsonify({'status': 'ok', 'schedule': schedule})

@app.route('/scan', methods=['GET','POST'])
@login_required
def scan():
    parsed, conf, raw, calc = None, None, None, None
    office_rate_percent = float(get_setting('office_interest_rate', '0.05')) * 100
    if request.method == 'POST':
        file = request.files.get('document')
        if file and file.filename:
            if not allowed_file(file.filename):
                flash('File harus berupa gambar (PNG, JPG, JPEG)', 'danger')
                return render_template('scan.html', parsed=parsed, confidence=conf, raw_text=raw, calculations=calc, office_rate_percent=office_rate_percent)
            path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(path)
            raw = do_ocr(path)
            try: os.remove(path)
            except: pass
            if raw:
                parsed, conf, raw = smart_parse(raw)
                if parsed: calc = calculate(parsed)
    return render_template('scan.html', parsed=parsed, confidence=conf, raw_text=raw, calculations=calc, office_rate_percent=office_rate_percent)

@app.route('/form')
@login_required
def online_form():
    office_rate_percent = float(get_setting('office_interest_rate', '0.05')) * 100
    return render_template('cabang_form.html', office_rate_percent=office_rate_percent)

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if session.get('role') != 'hr':
        flash('Hanya HR yang bisa mengakses halaman ini', 'danger')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        raw_pct = request.form.get('office_rate_percent', '').strip().replace(',', '.')
        try:
            pct = float(raw_pct)
            if pct <= 0 or pct > 100:
                flash('Persentase harus antara 0 dan 100', 'danger')
            else:
                set_setting('office_interest_rate', str(pct / 100), session.get('name'))
                flash(f'Suku bunga Non-Operasional diperbarui menjadi {pct}%', 'success')
        except ValueError:
            flash('Masukkan angka yang valid, cth. 8.5', 'danger')
        return redirect(url_for('settings'))

    current_percent = float(get_setting('office_interest_rate', '0.05')) * 100
    audit = get_setting_full('office_interest_rate')
    return render_template('settings.html', current_percent=current_percent, audit=audit)

@app.route('/calculator')
@login_required
def calculator(): return render_template('calculator.html')

@app.route('/api/submit', methods=['POST'])
@login_required
def submit():
    try:
        data = request.json
        calc = calculate(data)
        if not calc: return jsonify({'status':'error'}), 400
        data.update(calc)
        application_id = f"TAF-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
        conn = get_db()
        conn.execute('''INSERT INTO submissions 
            (application_id,npk,nama_lengkap,employee_type,jabatan_gol,departemen_cabang,cabang,
             type_motor,no_rangka,no_mesin,bpkb,
             loan_amount,down_payment,tenure_months,interest_rate,tanggal_mulai,
             principal,total_interest,monthly_installment,outstanding_balance,
             remarks,signature,email,phone,submitted_by,document_source)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (application_id,data.get('npk'),data.get('nama_lengkap'),data.get('employee_type','field'),
             data.get('jabatan_gol'),data.get('departemen_cabang'),data.get('cabang',session.get('branch')),
             data.get('type_motor'),data.get('no_rangka'),data.get('no_mesin'),data.get('bpkb'),
             data.get('loan_amount',16000000),data.get('down_payment',0),
             data.get('tenure_months',36),data.get('interest_rate',0.05),data.get('tanggal_mulai'),
             data.get('principal'),data.get('total_interest'),
             data.get('monthly_installment'),data.get('outstanding_balance'),
             data.get('remarks',''),data.get('signature',''),
             data.get('email',''),data.get('phone',''),
             session.get('name'),data.get('document_source','Manual')))
        conn.execute('''INSERT INTO employee_loans (npk, nama_lengkap, email, phone, active_loan_count, total_loans_ever, last_loan_date)
                      VALUES (?,?,?,?,1,1,?)
                      ON CONFLICT (npk) DO UPDATE SET
                          nama_lengkap = excluded.nama_lengkap,
                          email = excluded.email,
                          phone = excluded.phone,
                          active_loan_count = employee_loans.active_loan_count + 1,
                          total_loans_ever = employee_loans.total_loans_ever + 1,
                          last_loan_date = excluded.last_loan_date''',
                   (data.get('npk'),data.get('nama_lengkap'),data.get('email',''),data.get('phone',''),datetime.now()))
        conn.commit(); conn.close()
        return jsonify({'status':'success','application_id':application_id,'calculations':calc})
    except Exception as e:
        return jsonify({'status':'error','message':str(e)}),500

@app.route('/api/approve/<int:id>', methods=['POST'])
@login_required
def approve(id):
    if session.get('role') not in ['hr','manager']: return jsonify({'status':'error'}), 403
    body = request.get_json(silent=True) or {}
    conn = get_db()
    row = conn.execute("SELECT * FROM submissions WHERE id=?", (id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'status':'error','message':'Not found'}), 404
    sub = dict(row)

    loan_amount = float(body.get('loan_amount') if body.get('loan_amount') not in (None,'') else (sub.get('loan_amount') or 16000000))
    down_payment = float(body.get('down_payment') if body.get('down_payment') not in (None,'') else (sub.get('down_payment') or 0))
    tenure_months = int(body.get('tenure_months') if body.get('tenure_months') not in (None,'') else (sub.get('tenure_months') or 36))
    insurance_raw = body.get('insurance_amount')
    if insurance_raw in (None, ''):
        insurance_amount = sub.get('insurance_amount')
        if insurance_amount is None:
            insurance_amount = round(loan_amount * 0.053, 3)
    else:
        insurance_amount = float(insurance_raw)

    calc = calculate({'loan_amount': loan_amount, 'down_payment': down_payment,
                       'tenure_months': tenure_months, 'employee_type': sub.get('employee_type')})
    if not calc:
        conn.close()
        return jsonify({'status':'error','message':'Angka pinjaman tidak valid'}), 400

    conn.execute("UPDATE submissions SET status_approval='Approved', approved_date=CURRENT_TIMESTAMP, "
                 "loan_amount=?, down_payment=?, tenure_months=?, insurance_amount=?, "
                 "principal=?, total_interest=?, monthly_installment=?, outstanding_balance=?, interest_rate=? "
                 "WHERE id=?",
                 (loan_amount, down_payment, tenure_months, insurance_amount,
                  calc['principal'], calc['total_interest'], calc['monthly_installment'],
                  calc['outstanding_balance'], calc['interest_rate'], id))

    conn.execute("DELETE FROM loan_schedule WHERE submission_id=?", (id,))
    start_raw = sub.get('tanggal_mulai') or ''
    try:
        start_date = datetime.strptime(start_raw[:10], '%Y-%m-%d').date()
    except ValueError:
        start_date = datetime.now().date()

    balance = calc['outstanding_balance']
    monthly = calc['monthly_installment']
    for month in range(1, tenure_months + 1):
        due = add_months(start_date, month)
        balance = round(balance - monthly, 2)
        conn.execute("INSERT INTO loan_schedule (submission_id, month_number, due_date, expected_amount, balance_after) "
                     "VALUES (?,?,?,?,?)", (id, month, due.isoformat(), monthly, max(balance, 0)))

    conn.commit(); conn.close()
    return jsonify({'status':'ok'})

@app.route('/api/schedule/mark-paid/<int:schedule_id>', methods=['POST'])
@login_required
def mark_paid(schedule_id):
    if session.get('role') not in ['hr','manager']: return jsonify({'status':'error'}), 403
    body = request.get_json(silent=True) or {}
    paid = bool(body.get('paid', True))
    conn = get_db()
    if paid:
        conn.execute("UPDATE loan_schedule SET is_paid=1, paid_date=CURRENT_TIMESTAMP WHERE id=?", (schedule_id,))
    else:
        conn.execute("UPDATE loan_schedule SET is_paid=0, paid_date=NULL WHERE id=?", (schedule_id,))
    conn.commit(); conn.close()
    return jsonify({'status':'ok'})

@app.route('/reports/schedule')
@login_required
def schedule_report():
    if session.get('role') not in ['hr','manager']:
        return redirect(url_for('dashboard'))
    conn = get_db()
    rows = conn.execute(
        "SELECT ls.*, s.nama_lengkap, s.npk, s.cabang, s.type_motor "
        "FROM loan_schedule ls JOIN submissions s ON s.id = ls.submission_id "
        "WHERE s.status_approval = 'Approved' "
        "ORDER BY s.nama_lengkap, ls.month_number"
    ).fetchall()
    conn.close()

    today = datetime.now().date()
    loans = {}
    for r in rows:
        r = dict(r)
        sid = r['submission_id']
        if sid not in loans:
            loans[sid] = {'nama_lengkap': r['nama_lengkap'], 'npk': r['npk'],
                          'cabang': r['cabang'], 'type_motor': r['type_motor'], 'rows': []}
        due = datetime.strptime(r['due_date'], '%Y-%m-%d').date()
        if r['is_paid']:
            r['live_status'] = 'Paid'
        elif due < today:
            r['live_status'] = 'Overdue'
        elif due.year == today.year and due.month == today.month:
            r['live_status'] = 'Due this month'
        else:
            r['live_status'] = 'Upcoming'
        r['due_date_display'] = due.strftime('%d %b %Y')
        loans[sid]['rows'].append(r)

    return render_template('schedule_report.html', loans=list(loans.values()), today=today.strftime('%d %B %Y'))

@app.route('/export-schedule-excel', methods=['POST'])
@login_required
def export_schedule_excel():
    if session.get('role') not in ['hr','manager']:
        return redirect(url_for('dashboard'))
    included_ids = request.form.getlist('include')
    if not included_ids:
        flash('Tidak ada baris yang dipilih untuk diekspor', 'danger')
        return redirect(url_for('schedule_report'))

    placeholders = ','.join('?' * len(included_ids))
    conn = get_db()
    rows = conn.execute(
        f"SELECT ls.*, s.nama_lengkap, s.npk, s.cabang, s.type_motor "
        f"FROM loan_schedule ls JOIN submissions s ON s.id = ls.submission_id "
        f"WHERE ls.id IN ({placeholders}) "
        f"ORDER BY s.nama_lengkap, ls.month_number", included_ids
    ).fetchall()
    conn.close()

    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Jadwal Angsuran"
    headers = ['Nama', 'NPK', 'Cabang', 'Bulan', 'Jatuh Tempo', 'Jumlah', 'Sisa', 'Status']
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i + 1, value=h).font = Font(bold=True)

    today = datetime.now().date()
    for r_idx, row in enumerate(rows, start=2):
        r = dict(row)
        due = datetime.strptime(r['due_date'], '%Y-%m-%d').date()
        if r['is_paid']: status = 'Lunas'
        elif due < today: status = 'Terlambat'
        elif due.year == today.year and due.month == today.month: status = 'Bulan Ini'
        else: status = 'Akan Datang'
        ws.cell(row=r_idx, column=1, value=r['nama_lengkap'])
        ws.cell(row=r_idx, column=2, value=r['npk'])
        ws.cell(row=r_idx, column=3, value=r['cabang'])
        ws.cell(row=r_idx, column=4, value=r['month_number'])
        ws.cell(row=r_idx, column=5, value=due.strftime('%d %b %Y'))
        ws.cell(row=r_idx, column=6, value=r['expected_amount']).number_format = '#,##0'
        ws.cell(row=r_idx, column=7, value=r['balance_after']).number_format = '#,##0'
        ws.cell(row=r_idx, column=8, value=status)
    for col, w in zip('ABCDEFGH', [22, 12, 14, 8, 14, 14, 14, 12]):
        ws.column_dimensions[col].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'Jadwal_Angsuran_{datetime.now().strftime("%Y%m%d")}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/reject/<int:id>', methods=['POST'])
@login_required
def reject(id):
    if session.get('role') not in ['hr','manager']: return jsonify({'status':'error'}), 403
    body = request.get_json(silent=True) or {}
    reason = body.get('reason', '')
    conn = get_db()
    conn.execute("UPDATE submissions SET status_approval='Rejected', rejection_reason=? WHERE id=?",(reason, id))
    conn.commit(); conn.close()
    return jsonify({'status':'ok'})

@app.route('/export-excel')
@login_required
def export():
    if session.get('role') not in ['hr','manager']: return redirect(url_for('dashboard'))
    return send_file(generate_excel('full'), download_name=f'Report_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/export-excel/approved')
@login_required
def export_approved():
    if session.get('role') not in ['hr','manager']: return redirect(url_for('dashboard'))
    return send_file(generate_excel('approved'), download_name=f'Approved_{datetime.now().strftime("%Y%m%d")}.xlsx')

# Accounting-style number formats, copied exactly from the reference workbook
_ACC_FMT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
_ACC_FMT_JOURNAL = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
_PCT_FMT_2 = '0.00%'
_PCT_FMT_3 = '0.000%'

def _id_number(n):
    """Format a number Indonesian-style: 1.234.567 (dot thousands separator)."""
    return f"{n:,.0f}".replace(",", ".")

def build_edlin_sheet(ws, sub):
    """Write one submission's EDLIN journal into the given worksheet, matching
    TAF's reference template cell-for-cell AND visually: same rows/columns, same
    live Excel formulas, same journal entries, same fonts/colors/borders/widths.
    Shared by both the single-submission export and the per-employee export."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    otr = float(sub.get('loan_amount') or 24389653)
    dp = float(sub.get('down_payment') or 6000000)
    tenor = int(sub.get('tenure_months') or 24)
    rate = 0.069  # COF + 1% -- TAF's internal EDLIN funding rate, separate from
                  # the employee-facing 5% flat rate used elsewhere in the app.
    saved_insurance = sub.get('insurance_amount')
    insurance = float(saved_insurance) if saved_insurance is not None else round(otr * 0.053, 3)

    monthly_rate = rate / 12
    total_principal = (otr - dp) + insurance
    monthly_installment = round(total_principal * monthly_rate * (1 + monthly_rate) ** tenor / ((1 + monthly_rate) ** tenor - 1))
    total_ar = monthly_installment * tenor

    bold = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    journal_header_font = Font(bold=True, color='FF0070C0')
    note_font = Font(color='FF0070C0')
    thin = Side(style='thin')
    all_thin = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_center_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.column_dimensions['A'].width = 22.27
    ws.column_dimensions['B'].width = 13.73
    ws.column_dimensions['C'].width = 14.63
    ws.column_dimensions['D'].width = 16.73
    ws.column_dimensions['E'].width = 14.27
    ws.column_dimensions['N'].width = 11.82

    ws['H1'] = 'Skema Jurnal EDLIN'
    ws['H1'].font = title_font
    ws['H1'].border = Border(bottom=Side(style='medium'))
    ws['A2'] = 'Nama Karyawan'; ws['A2'].alignment = Alignment(horizontal='left')
    ws['B2'] = sub.get('nama_lengkap', '')
    ws['A3'] = 'NPK'; ws['A3'].alignment = Alignment(horizontal='left')
    ws['B3'] = sub.get('npk', '')

    for coord, label, val, fmt in [
        ('5', 'OTR', otr, _ACC_FMT), ('6', 'DP (min 10%)', dp, _ACC_FMT), ('7', 'Sisa', '=B5-B6', _ACC_FMT),
        ('8', 'Insurance', insurance, _ACC_FMT), ('9', 'Total Principal', '=SUM(B7:B8)', _ACC_FMT),
        ('10', 'Interest', '=B11-B9', _ACC_FMT), ('11', 'Total AR', '=B16*B14', _ACC_FMT),
        ('13', 'Rate', rate, _PCT_FMT_2), ('14', 'Tenor', tenor, None),
        ('15', 'Bunga per bulan', '=B13/12', _PCT_FMT_3), ('16', 'Angsuran total anuitas', '=ROUND(-PMT(B15,B14,B9),0)', _ACC_FMT),
    ]:
        ws[f'A{coord}'] = label
        ws[f'A{coord}'].alignment = Alignment(horizontal='left')
        ws[f'B{coord}'] = val
        if fmt: ws[f'B{coord}'].number_format = fmt

    ws['C8'] = 'Total Premi'
    ws['C10'] = 'Effective Rate'
    ws['C13'] = 'Effective Rate'

    ws['B9'].font = bold
    ws['B9'].border = Border(top=thin)

    ws['H3'] = 'Jurnal pembentukan Edlin dan pembayaran ke Dealer'; ws['H3'].font = journal_header_font
    ws['H4'] = '1213101301 - CONS FIN - EDLIN'
    ws['N4'] = '=B11-B8'; ws['N4'].number_format = _ACC_FMT_JOURNAL
    ws['O4'] = 0; ws['O4'].number_format = _ACC_FMT
    ws['Q4'] = f'Total Consfin Edlin menjadi sama dengan yang harus dibayar Karyawan = {_id_number(total_ar)}'
    ws['Q4'].font = note_font
    ws['Q4'].alignment = center_wrap
    ws.merge_cells('Q4:R9')
    ws['H5'] = '1223101301 - CONS FIN - UNEARNED - AR EDLIN'
    ws['N5'] = 0; ws['N5'].number_format = _ACC_FMT_JOURNAL
    ws['O5'] = '=B10'; ws['O5'].number_format = _ACC_FMT
    ws['H6'] = '2222100101 - AP SUPPLIER (DEALER)'
    ws['N6'] = 0; ws['N6'].number_format = _ACC_FMT_JOURNAL
    ws['O6'] = '=B7'; ws['O6'].number_format = _ACC_FMT

    ws['H8'] = 'Jurnal pembayaran insurance ke AAB'; ws['H8'].font = journal_header_font
    ws['H9'] = '1213101301 - CONS FIN - EDLIN'
    ws['N9'] = '=B8'; ws['N9'].number_format = _ACC_FMT_JOURNAL
    ws['O9'] = 0; ws['O9'].number_format = _ACC_FMT_JOURNAL
    ws['H10'] = '2222100101 - AP SUPPLIER (AAB)'
    ws['N10'] = 0; ws['N10'].number_format = _ACC_FMT_JOURNAL
    ws['O10'] = '=B8'; ws['O10'].number_format = _ACC_FMT_JOURNAL

    ws['H12'] = (f'Jurnal pemotongan payroll karyawan (tiap bulan {_id_number(monthly_installment)} '
                 f'sampai {tenor} bulan = {_id_number(total_ar)})')
    ws['H12'].font = journal_header_font
    ws['H13'] = '7011100101 - SALARIES'
    ws['N13'] = '=B16*B14'; ws['N13'].number_format = _ACC_FMT_JOURNAL
    ws['O13'] = 0; ws['O13'].number_format = _ACC_FMT_JOURNAL
    ws['H14'] = '1941100801 - LTE - MOTORCYCLE (KPM)'
    ws['N14'] = 0; ws['N14'].number_format = _ACC_FMT_JOURNAL
    ws['O14'] = '=B16*B14'; ws['O14'].number_format = _ACC_FMT_JOURNAL

    ws['H16'] = 'Jurnal pembalikan LTE KPM ke Consfin Edlin'; ws['H16'].font = journal_header_font
    ws['H17'] = '1941100801 - LTE - MOTORCYCLE (KPM)'
    ws['N17'] = '=O14'; ws['N17'].number_format = _ACC_FMT_JOURNAL
    ws['H18'] = '1213101301 - CONS FIN - EDLIN'
    ws['O18'] = '=N17'; ws['O18'].number_format = _ACC_FMT_JOURNAL

    headers = ['Month', 'Principal', 'Interest', 'Installment', 'Balance']
    for i, label in enumerate(headers):
        col = 'ABCDE'[i]
        cell = ws[f'{col}19']
        cell.value = label
        cell.font = bold
        cell.fill = header_fill
        cell.border = all_thin
        cell.alignment = left_center_wrap if col == 'A' else center_wrap

    ws['H20'] = 'Jurnal pengakuan income atas Interest Edlin'; ws['H20'].font = journal_header_font
    ws['H21'] = '1223101301 - CONS FIN - UNEARNED - AR EDLIN'
    ws['N21'] = '=O5'; ws['N21'].number_format = _ACC_FMT_JOURNAL
    ws['O21'] = 0; ws['O21'].number_format = _ACC_FMT_JOURNAL
    ws['H22'] = '4991199901 - OTHER INCOME - OTHERS'
    ws['N22'] = 0; ws['N22'].number_format = _ACC_FMT_JOURNAL
    ws['O22'] = '=O5'; ws['O22'].number_format = _ACC_FMT_JOURNAL

    ws['A20'] = 0
    ws['A20'].border = all_thin
    ws['A20'].alignment = center_wrap
    ws['E20'] = '=B9'; ws['E20'].number_format = _ACC_FMT
    ws['E20'].border = all_thin
    for col in 'BCD':
        ws[f'{col}20'].border = all_thin

    for month in range(1, tenor + 1):
        r = 20 + month
        ws[f'A{r}'] = month
        ws[f'B{r}'] = f'=D{r}-C{r}'; ws[f'B{r}'].number_format = _ACC_FMT
        ws[f'C{r}'] = f'=E{r-1}*$B$15'; ws[f'C{r}'].number_format = _ACC_FMT
        ws[f'D{r}'] = '=B16' if month == 1 else f'=D{r-1}'
        ws[f'D{r}'].number_format = _ACC_FMT
        ws[f'E{r}'] = f'=ROUND(E{r-1}-B{r},0)'; ws[f'E{r}'].number_format = _ACC_FMT
        for col in 'ABCDE':
            ws[f'{col}{r}'].border = all_thin
            ws[f'{col}{r}'].alignment = center_wrap if col == 'A' else Alignment(vertical='center', wrap_text=True)


def unique_sheet_name(base_name, used_names):
    """Sanitize into a valid, unique Excel sheet name (max 31 chars, no : \\ / ? * [ ])."""
    safe = re.sub(r'[:\\/?*\[\]]', '-', str(base_name))[:31]
    name = safe
    n = 1
    while name in used_names:
        n += 1
        suffix = f"_{n}"
        name = f"{safe[:31-len(suffix)]}{suffix}"
    return name


def compute_monthly_journal_entries(target_year, target_month):
    """For a given calendar month, find every EDLIN journal line that should be
    booked that period, across ALL approved loans, regardless of when each loan
    individually started."""
    conn = get_db()
    subs = [dict(r) for r in conn.execute("SELECT * FROM submissions WHERE status_approval='Approved'").fetchall()]
    conn.close()

    entries = []
    for sub in subs:
        otr = float(sub.get('loan_amount') or 0)
        dp = float(sub.get('down_payment') or 0)
        tenor = int(sub.get('tenure_months') or 0)
        if not otr or not tenor:
            continue
        saved_insurance = sub.get('insurance_amount')
        insurance = float(saved_insurance) if saved_insurance is not None else round(otr * 0.053, 3)
        rate = 0.069

        start_raw = sub.get('tanggal_mulai') or ''
        try:
            start_date = datetime.strptime(start_raw[:10], '%Y-%m-%d').date()
        except ValueError:
            continue

        name = sub.get('nama_lengkap', '') or '-'
        npk = sub.get('npk', '') or '-'

        sisa = otr - dp
        total_principal = sisa + insurance
        monthly_rate = rate / 12
        installment = round(total_principal * monthly_rate * (1 + monthly_rate) ** tenor / ((1 + monthly_rate) ** tenor - 1))
        total_ar = installment * tenor
        total_interest_lifetime = total_ar - total_principal

        if start_date.year == target_year and start_date.month == target_month:
            grp = 'Pembentukan Edlin & Pembayaran Dealer'
            entries.append({'jenis': grp, 'akun': '1213101301 - CONS FIN - EDLIN', 'debit': total_ar - insurance, 'kredit': 0, 'nama': name, 'npk': npk})
            entries.append({'jenis': grp, 'akun': '1223101301 - CONS FIN - UNEARNED - AR EDLIN', 'debit': 0, 'kredit': total_interest_lifetime, 'nama': name, 'npk': npk})
            entries.append({'jenis': grp, 'akun': '2222100101 - AP SUPPLIER (DEALER)', 'debit': 0, 'kredit': sisa, 'nama': name, 'npk': npk})
            grp = 'Pembayaran Insurance ke AAB'
            entries.append({'jenis': grp, 'akun': '1213101301 - CONS FIN - EDLIN', 'debit': insurance, 'kredit': 0, 'nama': name, 'npk': npk})
            entries.append({'jenis': grp, 'akun': '2222100101 - AP SUPPLIER (AAB)', 'debit': 0, 'kredit': insurance, 'nama': name, 'npk': npk})
            grp = 'Pembalikan LTE KPM ke Consfin Edlin'
            entries.append({'jenis': grp, 'akun': '1941100801 - LTE - MOTORCYCLE (KPM)', 'debit': total_ar, 'kredit': 0, 'nama': name, 'npk': npk})
            entries.append({'jenis': grp, 'akun': '1213101301 - CONS FIN - EDLIN', 'debit': 0, 'kredit': total_ar, 'nama': name, 'npk': npk})

        balance = total_principal
        for m in range(1, tenor + 1):
            due = add_months(start_date, m)
            interest_amt = balance * monthly_rate
            principal_amt = installment - interest_amt
            balance = round(balance - principal_amt, 0)
            if due.year == target_year and due.month == target_month:
                grp = f'Pemotongan Payroll (Bulan ke-{m})'
                entries.append({'jenis': grp, 'akun': '7011100101 - SALARIES', 'debit': installment, 'kredit': 0, 'nama': name, 'npk': npk})
                entries.append({'jenis': grp, 'akun': '1941100801 - LTE - MOTORCYCLE (KPM)', 'debit': 0, 'kredit': installment, 'nama': name, 'npk': npk})
                grp = f'Pengakuan Income Interest (Bulan ke-{m})'
                entries.append({'jenis': grp, 'akun': '1223101301 - CONS FIN - UNEARNED - AR EDLIN', 'debit': round(interest_amt), 'kredit': 0, 'nama': name, 'npk': npk})
                entries.append({'jenis': grp, 'akun': '4991199901 - OTHER INCOME - OTHERS', 'debit': 0, 'kredit': round(interest_amt), 'nama': name, 'npk': npk})
                break

    entries.sort(key=lambda e: (e['jenis'], e['nama']))
    return entries


@app.route('/reports/edlin-monthly')
@login_required
def edlin_monthly_report():
    if session.get('role') not in ['hr', 'manager']:
        return redirect(url_for('dashboard'))
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    entries = compute_monthly_journal_entries(year, month) if (year and month) else []
    total_debit = sum(e['debit'] for e in entries)
    total_kredit = sum(e['kredit'] for e in entries)
    return render_template('edlin_monthly.html', entries=entries, year=year, month=month,
                            total_debit=total_debit, total_kredit=total_kredit,
                            current_year=datetime.now().year)


@app.route('/export-edlin-monthly')
@login_required
def export_edlin_monthly():
    if session.get('role') not in ['hr', 'manager']:
        return redirect(url_for('dashboard'))
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not year or not month:
        flash('Pilih bulan dan tahun terlebih dahulu', 'danger')
        return redirect(url_for('edlin_monthly_report'))

    entries = compute_monthly_journal_entries(year, month)

    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = f'Jurnal {month:02d}-{year}'
    headers = ['Jenis Jurnal', 'Akun', 'Debit', 'Kredit', 'Nama Karyawan', 'NPK']
    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=i + 1, value=h)
        cell.font = Font(bold=True)
    for r, e in enumerate(entries, start=2):
        ws.cell(row=r, column=1, value=e['jenis'])
        ws.cell(row=r, column=2, value=e['akun'])
        ws.cell(row=r, column=3, value=e['debit']).number_format = '#,##0'
        ws.cell(row=r, column=4, value=e['kredit']).number_format = '#,##0'
        ws.cell(row=r, column=5, value=e['nama'])
        ws.cell(row=r, column=6, value=e['npk'])
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['E'].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'Jurnal_Bulanan_{year}-{month:02d}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export-edlin/batch/<status>')
@login_required
def export_edlin_batch(status):
    if session.get('role') not in ['hr', 'manager']:
        return redirect(url_for('dashboard'))
    if status not in ('Pending', 'Approved'):
        flash('Status tidak valid', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db()
    subs = [dict(r) for r in conn.execute(
        "SELECT * FROM submissions WHERE status_approval=? ORDER BY nama_lengkap, submitted_date", (status,)
    ).fetchall()]
    conn.close()

    if not subs:
        flash(f'Tidak ada pengajuan dengan status {status}', 'danger')
        return redirect(url_for('dashboard'))

    from openpyxl import Workbook
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.remove(wb.active)

    used_names = set()
    for sub in subs:
        base_name = re.sub(r'[:\\/?*\[\]]', '-', f"{(sub.get('nama_lengkap') or 'NoName')[:18]}_{sub['id']}")[:31]
        name = base_name
        n = 1
        while name in used_names:
            n += 1
            name = f"{base_name[:28]}_{n}"
        used_names.add(name)
        ws = wb.create_sheet(title=name)
        build_edlin_sheet(ws, sub)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'EDLIN_{status}_{datetime.now().strftime("%Y%m%d")}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export-edlin/<int:id>')
@login_required
def export_edlin(id):
    """Generate EDLIN-formatted Excel matching their exact template"""
    if session.get('role') not in ['hr','manager']:
        return redirect(url_for('dashboard'))

    conn = get_db()
    sub = dict(conn.execute("SELECT * FROM submissions WHERE id=?", (id,)).fetchone())
    conn.close()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EDLIN Calculation"
    build_edlin_sheet(ws, sub)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=f'EDLIN_{sub.get("npk","")}_{sub.get("nama_lengkap","")}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export-edlin-employee/<npk>')
@login_required
def export_edlin_employee(npk):
    """Generate one EDLIN-formatted workbook with a sheet per submission for this NPK's full loan history"""
    if session.get('role') not in ['hr','manager']:
        return redirect(url_for('dashboard'))

    conn = get_db()
    subs = [dict(r) for r in conn.execute(
        "SELECT * FROM submissions WHERE npk=? ORDER BY submitted_date ASC", (npk,)).fetchall()]
    conn.close()

    if not subs:
        flash('Tidak ada pengajuan untuk NPK ini', 'danger')
        return redirect(url_for('dashboard'))

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    for i, sub in enumerate(subs, 1):
        name = unique_sheet_name(f"{i}. {sub.get('application_id') or sub.get('id')}", used_names)
        used_names.add(name)
        ws = wb.create_sheet(title=name)
        build_edlin_sheet(ws, sub)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nama = subs[-1].get('nama_lengkap', '')
    return send_file(output, download_name=f'EDLIN_{npk}_{nama}_History.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/')
def index():
    if 'user' in session: return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=5000)
